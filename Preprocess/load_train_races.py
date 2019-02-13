import openpyxl as op
import os
import pickle


home = os.path.dirname(os.getcwd())

wb = op.load_workbook(home + '\\Dataset\\demographic-others-labels.xlsx')
ws = wb[wb.sheetnames[1]]

races = []
img_names = []

for row in range(2, ws.max_row+1):
    img_names.append(ws.cell(row=row, column=1).value)
    races.append(ws.cell(row=row, column=19).value)

svfile_races = open(home + '\\Attributes\\races.dat', mode='wb')
svfile_imgs = open(home + '\\Attributes\\imgs.dat', mode='wb')

pickle.dump(races, svfile_races)
pickle.dump(img_names, svfile_imgs)

svfile_races.close()
svfile_imgs.close()
